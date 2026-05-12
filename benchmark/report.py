"""Sonuç raporlama: CSV, formatlı XLSX, per-SNR pivot ve metrik grafikleri.

Hem bench_real (tek wav, kalite metriği yok) hem bench_synthetic (model x snr x pair)
çıktılarını destekler. Fonksiyonlar eksik kolonlara karşı toleranslıdır: tablo neyi
verirse onu yazar.
"""

from __future__ import annotations

import csv
from typing import Iterable


# Tüm olası kolonlar için Türkçe başlık eşlemesi. Eksik olanlar dokunulmaz geçer.
_RENAME = {
    "model": "Model",
    "load_time_s": "Yükleme (s)",
    "process_time_s": "İşlem (s)",
    "process_time_mean_s": "İşlem Ort. (s)",
    "process_time_std_s": "İşlem Std (s)",
    "audio_duration_s": "Ses Süresi (s)",
    "rtf": "RTF",
    "rtf_mean": "RTF Ort.",
    "rtf_std": "RTF Std",
    "n_repeats": "Tekrar",
    "ram_load_mb": "RAM Yükleme (MB)",
    "ram_process_mb": "RAM İşlem (MB)",
    "peak_ram_mb": "Peak RAM (MB)",
    "model_size_mb": "Boyut (MB)",
    "param_count": "Parametre Sayısı",
    "si_sdr": "SI-SDR (dB)",
    "stoi": "STOI",
    "pesq": "PESQ",
    "snr_db": "SNR (dB)",
    "clean_file": "Clean",
    "noise_file": "Noise",
}


def save_csv(results: list[dict], path: str) -> None:
    """Sonuçları CSV olarak kaydet.

    Tüm satırlarda görülen kolonların birleşimini kullanır; eksik alanlar boş.
    """
    if not results:
        return
    fieldnames: list[str] = []
    seen = set()
    for r in results:
        for k in r.keys():
            if k not in seen:
                seen.add(k)
                fieldnames.append(k)
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(results)


def save_xlsx(results: list[dict], path: str, sheet_name: str = "Benchmark") -> None:
    """Sonuçları formatlı bir Excel dosyası olarak kaydet.
    Başlıklar koyu/dolgulu, ilk kolon sola diğerleri sağa yaslı, kolon genişlikleri otomatik.
    """
    if not results:
        return

    import pandas as pd
    from openpyxl.styles import Font, Alignment, PatternFill

    df = pd.DataFrame(results)

    # Bilinen kolonları Türkçe başlığa çevir
    df = df.rename(columns={k: v for k, v in _RENAME.items() if k in df.columns})

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        ws = writer.sheets[sheet_name]

        # Başlık satırını formatla (1. satır)
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4472C4")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Kolon genişliklerini otomatik ayarla
        for col in ws.columns:
            max_len = max(
                (len(str(cell.value)) for cell in col if cell.value is not None),
                default=10,
            )
            ws.column_dimensions[col[0].column_letter].width = max_len + 3

        # Veri satırları: ilk kolon sola, diğerleri sağa
        for row in ws.iter_rows(min_row=2):
            for i, cell in enumerate(row):
                cell.alignment = Alignment(horizontal="left" if i == 0 else "right")


def print_table(results: Iterable[dict]) -> None:
    """Sonuçları terminalde okunabilir bir tablo olarak bas (eski formatla uyumlu)."""
    results = list(results)
    if not results:
        return

    # Mevcut anahtarlara göre dinamik kolon seçimi
    candidates = [
        ("model", "Model"),
        ("process_time_mean_s", "İşlem(s)"),
        ("process_time_s", "İşlem(s)"),
        ("rtf_mean", "RTF"),
        ("rtf", "RTF"),
        ("load_time_s", "Yükleme(s)"),
        ("peak_ram_mb", "Peak RAM(MB)"),
        ("ram_process_mb", "RAM(MB)"),
        ("model_size_mb", "Boyut(MB)"),
        ("param_count", "Parametre"),
        ("si_sdr", "SI-SDR"),
        ("stoi", "STOI"),
        ("pesq", "PESQ"),
    ]
    cols: list[tuple[str, str]] = []
    seen_titles = set()
    available = set().union(*(set(r.keys()) for r in results))
    for key, title in candidates:
        if key in available and title not in seen_titles:
            cols.append((key, title))
            seen_titles.add(title)

    header = " | ".join(f"{title:>14}" for _, title in cols)
    print("\n" + header)
    print("-" * len(header))
    for r in results:
        row = " | ".join(f"{str(r.get(key, '')):>14}" for key, _ in cols)
        print(row)


# ---------------------------------------------------------------------------
# Sentetik benchmark: per-SNR pivot ve grafikler
# ---------------------------------------------------------------------------

# Pivot ve grafiklerde kullanılacak metrikler. Eksik olanlar atlanır.
_AGG_METRICS = [
    ("rtf_mean", "RTF"),
    ("si_sdr", "SI-SDR"),
    ("stoi", "STOI"),
    ("pesq", "PESQ"),
    ("peak_ram_mb", "Peak RAM (MB)"),
]


def save_per_snr_summary(rows: list[dict], path: str) -> None:
    """rows: her satır bir (model, snr, pair) deneyi.

    Her metrik için ayrı sheet'te pivot tablo: rows=model, cols=snr, value=mean.
    """
    if not rows:
        return

    import pandas as pd
    from openpyxl.styles import Font, Alignment, PatternFill

    df = pd.DataFrame(rows)
    if "model" not in df.columns or "snr_db" not in df.columns:
        # Yanlış format -- yine de raw'ı yaz
        save_xlsx(rows, path, sheet_name="Raw")
        return

    available_metrics = [(k, label) for k, label in _AGG_METRICS if k in df.columns]
    if not available_metrics:
        save_xlsx(rows, path, sheet_name="Raw")
        return

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for metric_key, metric_label in available_metrics:
            pivot = (
                df.pivot_table(
                    index="model",
                    columns="snr_db",
                    values=metric_key,
                    aggfunc="mean",
                )
                .round(3)
                .sort_index()
            )
            sheet = metric_label[:31]  # XLSX sheet ismi max 31 karakter
            pivot.to_excel(writer, sheet_name=sheet)

            ws = writer.sheets[sheet]
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill("solid", fgColor="4472C4")
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            for col in ws.columns:
                max_len = max(
                    (len(str(cell.value)) for cell in col if cell.value is not None),
                    default=10,
                )
                ws.column_dimensions[col[0].column_letter].width = max_len + 3


def plot_per_snr(rows: list[dict], metric: str, path: str, ylabel: str | None = None) -> None:
    """matplotlib line chart: x=SNR, y=mean(metric), her model bir çizgi."""
    if not rows:
        return

    import pandas as pd
    import matplotlib

    matplotlib.use("Agg")  # GUI yok, sadece PNG kaydı
    import matplotlib.pyplot as plt

    df = pd.DataFrame(rows)
    if metric not in df.columns or "model" not in df.columns or "snr_db" not in df.columns:
        return

    pivot = df.pivot_table(
        index="snr_db",
        columns="model",
        values=metric,
        aggfunc="mean",
    ).sort_index()

    # Tamamen NaN olan modelleri çizmenin anlamı yok (örn. pesq paketi yoksa)
    pivot = pivot.dropna(axis=1, how="all")
    if pivot.empty:
        return  # gerçek veri yoksa boş PNG üretme

    fig, ax = plt.subplots(figsize=(8, 5))
    for model_name in pivot.columns:
        ax.plot(pivot.index, pivot[model_name], marker="o", label=model_name)

    ax.set_xlabel("SNR (dB)")
    ax.set_ylabel(ylabel or metric)
    ax.set_title(f"{ylabel or metric} vs SNR")
    ax.grid(True, alpha=0.3)
    ax.legend(loc="best", fontsize=9)
    fig.tight_layout()
    fig.savefig(path, dpi=120)
    plt.close(fig)
