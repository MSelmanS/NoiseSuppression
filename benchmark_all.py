"""Tüm modelleri tek input üzerinde karşılaştıran benchmark scripti.

Ölçtükleri: yükleme süresi, işleme süresi, RTF, RAM farkı, model boyutu,
parametre sayısı. Her modelin çıktısını ayrı wav olarak kaydeder.
"""

import os
import csv
import gc
from datetime import datetime
import tkinter as tk
from tkinter import filedialog, messagebox

from audio_io.file_io import load_audio, save_audio
from benchmark.metrics import current_ram_mb, time_it, model_size_mb, param_count
from config import MODEL_SR

# Tüm modelleri import et
from models.spectral_subtraction import SpectralSubtraction
from models.rnnoise_model import RNNoiseDenoiser
from models.deepfilternet_model import DeepFilterNetDenoiser
from models.demucs_model import DemucsDns48, DemucsDns64, DemucsMaster64
from models.metricgan_model import MetricGANDenoiser


# Karşılaştırılacak modellerin listesi. Yeni model eklemek için buraya ekle.
MODELS = [
    SpectralSubtraction,
    RNNoiseDenoiser,
    DeepFilterNetDenoiser,
    DemucsDns48,
    DemucsDns64,
    DemucsMaster64,
    MetricGANDenoiser,
]


def pick_input_file():
    """Tkinter ile dosya seçimi. İptal edilirse None."""
    root = tk.Tk()
    root.withdraw()
    want_file = messagebox.askyesno(
        title="Benchmark girişi",
        message="Test için bir .wav dosyası seç.\n(İptal edersen çıkılır.)"
    )
    path = None
    if want_file:
        path = filedialog.askopenfilename(
            title=".wav dosyası seç",
            filetypes=[("WAV dosyaları", "*.wav"), ("Tüm dosyalar", "*.*")]
        )
        if not path:
            path = None
    root.destroy()
    return path


def run_one_benchmark(model_class, audio):
    """Tek bir model için tüm ölçümleri yap.
    Dönüş: (sonuç sözlüğü, temizlenmiş ses)"""

    # --- Yükleme ---
    ram_before = current_ram_mb()
    model = model_class()
    _, load_time = time_it(model.load)
    ram_after_load = current_ram_mb()

    # --- İşleme ---
    # Warmup yok çünkü gerçek kullanımdaki ilk çağrıyı ölçmek istiyoruz
    denoised, process_time = time_it(model.process, audio)
    ram_after_process = current_ram_mb()

    # --- Metrikler ---
    audio_duration = len(audio) / MODEL_SR
    rtf = process_time / audio_duration   # Real Time Factor, <1 ise real-time

    result = {
        "model": model.name,
        "load_time_s": round(load_time, 3),
        "process_time_s": round(process_time, 3),
        "audio_duration_s": round(audio_duration, 3),
        "rtf": round(rtf, 3),
        "ram_load_mb": round(ram_after_load - ram_before, 1),
        "ram_process_mb": round(ram_after_process - ram_after_load, 1),
        "model_size_mb": round(model_size_mb(model.model if hasattr(model, "model") else None), 2),
        "param_count": param_count(model.model if hasattr(model, "model") else None),
    }

    # Model nesnesini serbest bırak (bir sonraki model için RAM'i temizle)
    del model
    gc.collect()

    return result, denoised


def print_table(results):
    """Sonuçları terminalde okunabilir bir tablo olarak bas."""
    # Gösterilecek sütunlar ve başlıkları
    cols = [
        ("model", "Model"),
        ("process_time_s", "İşlem(s)"),
        ("rtf", "RTF"),
        ("load_time_s", "Yükleme(s)"),
        ("ram_process_mb", "RAM(MB)"),
        ("model_size_mb", "Boyut(MB)"),
        ("param_count", "Parametre"),
    ]

    # Başlıkları yazdır
    header = " | ".join(f"{title:>14}" for _, title in cols)
    print("\n" + header)
    print("-" * len(header))

    # Satırları yazdır
    for r in results:
        row = " | ".join(f"{str(r[key]):>14}" for key, _ in cols)
        print(row)


def save_csv(results, path):
    """Sonuçları CSV olarak kaydet."""
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=list(results[0].keys()))
        writer.writeheader()
        writer.writerows(results)


def save_xlsx(results, path):
    """Sonuçları formatlı bir Excel dosyası olarak kaydet.
    Başlıklar koyu, sayısal sütunlar sağa yaslı, kolon genişlikleri içeriğe göre."""
    import pandas as pd
    from openpyxl.styles import Font, Alignment, PatternFill

    df = pd.DataFrame(results)

    # Türkçe başlıklar - okunurluğu artırır
    rename = {
        "model": "Model",
        "load_time_s": "Yükleme (s)",
        "process_time_s": "İşlem (s)",
        "audio_duration_s": "Ses Süresi (s)",
        "rtf": "RTF",
        "ram_load_mb": "RAM Yükleme (MB)",
        "ram_process_mb": "RAM İşlem (MB)",
        "model_size_mb": "Boyut (MB)",
        "param_count": "Parametre Sayısı",
    }
    df = df.rename(columns=rename)

    # Önce düz yaz, sonra openpyxl ile stilleri ekleyeceğiz
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Benchmark", index=False)
        ws = writer.sheets["Benchmark"]

        # Başlık satırını formatla (1. satır)
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4472C4")  # koyu mavi
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Kolon genişliklerini otomatik ayarla (her kolonun en uzun içeriğine göre)
        for col in ws.columns:
            max_len = max(len(str(cell.value)) for cell in col if cell.value is not None)
            ws.column_dimensions[col[0].column_letter].width = max_len + 3

        # Veri satırlarını hizala: ilk kolon (Model) sola, diğerleri sağa
        for row in ws.iter_rows(min_row=2):
            for i, cell in enumerate(row):
                cell.alignment = Alignment(horizontal="left" if i == 0 else "right")


def main():
    # 1) Input
    path = pick_input_file()
    if path is None:
        print("Dosya seçilmedi, çıkılıyor.")
        return

    print(f"Giriş: {path}")
    audio, _ = load_audio(path)
    print(f"Ses uzunluğu: {len(audio) / MODEL_SR:.2f} sn, {len(audio)} sample\n")

    # 2) Zaman damgalı çıktı klasörü oluştur
    #    Format: output/output_YYYYMMDD_HHMMSS
    #    Her çalıştırma ayrı klasöre gider, üstüne yazılmaz.
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_dir = os.path.join("output", f"output_{timestamp}")
    os.makedirs(out_dir, exist_ok=True)
    print(f"Çıktı klasörü: {out_dir}\n")

    # Orijinal girişi de çıktı klasörüne kopyala (karşılaştırma için)
    save_audio(os.path.join(out_dir, "00_original.wav"), audio)

    # 3) Her modeli çalıştır
    results = []
    for i, model_class in enumerate(MODELS, start=1):
        print(f"[{i}/{len(MODELS)}] {model_class.__name__} çalıştırılıyor...")
        try:
            result, denoised = run_one_benchmark(model_class, audio)
            results.append(result)

            # Temizlenmiş sesi kaydet
            out_wav = os.path.join(out_dir, f"{i:02d}_{result['model']}.wav")
            save_audio(out_wav, denoised)
            print(f"    OK  -> RTF={result['rtf']}, süre={result['process_time_s']}s")
        except Exception as e:
            # Bir model hata verirse toplu çalışma ölmesin
            print(f"    HATA: {e}")

    # 4) Sonuçları göster ve kaydet
    if results:
        print_table(results)
        csv_path = os.path.join(out_dir, "results.csv")
        xlsx_path = os.path.join(out_dir, "results.xlsx")
        save_csv(results, csv_path)
        save_xlsx(results, xlsx_path)
        print(f"\nSonuçlar CSV : {csv_path}")
        print(f"Sonuçlar XLSX: {xlsx_path}")
        print(f"Temizlenmiş sesler: {out_dir}/")


if __name__ == "__main__":
    main()